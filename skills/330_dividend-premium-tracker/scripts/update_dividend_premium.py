#!/usr/bin/env python3
"""
股息率溢价跟踪 - 自动化更新脚本
数据来源：
- 股息率：中证指数官网 H30269 指数估值
- 10年期国债收益率：财政部官网
"""

import xlrd
import csv
import os
import sys
import subprocess
import re
from datetime import datetime

# 配置
DATA_DIR = "/Users/liyi/.openclaw/workspace"
CSV_FILE = os.path.join(DATA_DIR, "股息率溢价跟踪.csv")
EXCEL_FILE = os.path.join(DATA_DIR, "股息率溢价跟踪.xlsx")

def download_dividend_rate(date_str):
    """从中证指数官网下载股息率数据"""
    url = "https://oss-ch.csindex.com.cn/static/html/csindex/public/uploads/file/autofile/indicator/H30269indicator.xls"
    local_file = "/tmp/H30269indicator.xls"
    
    os.system(f"curl -s -o {local_file} '{url}'")
    
    try:
        book = xlrd.open_workbook(local_file)
        sheet = book.sheet_by_index(0)
        
        for row in range(1, sheet.nrows):
            row_date = str(sheet.cell_value(row, 0))
            if row_date == date_str:
                return sheet.cell_value(row, 8)  # 股息率1
    except Exception as e:
        print(f"下载股息率失败: {e}")
    
    return None

def get_bond_yield(date_str):
    """从财政部官网获取10年期国债收益率"""
    result = subprocess.run(
        ['curl', '-s', 'https://yield.chinabond.com.cn/cbweb-czb-web/czb/moreInfo?locale=cn_ZH&nameType=1'],
        capture_output=True, text=True
    )
    
    text = result.stdout
    
    # 查找对应日期的10年数据
    # 格式: 日期 ... 10年 ... 收益率
    pattern = rf'{date_str}.*?10年.*?(\d+\.\d{{2}})'
    match = re.search(pattern, text, re.DOTALL)
    
    if match:
        return float(match.group(1))
    
    return None

def load_existing_data():
    """加载现有数据"""
    data = []
    if os.path.exists(CSV_FILE):
        with open(CSV_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(row)
    return data

def save_data(data):
    """保存数据"""
    if not data:
        return
    
    keys = ['日期', '股息率1', '10年国债收益率', '股息率溢价']
    with open(CSV_FILE, 'w', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(data)

def generate_excel():
    """生成带图表的Excel"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    df = pd.read_csv(CSV_FILE)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, float):
                ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    # 图表
    chart = LineChart()
    chart.title = "股息率溢价走势"
    chart.style = 10
    chart.x_axis.title = "日期"
    chart.y_axis.title = "溢价 (%)"
    chart.width = 18
    chart.height = 10
    
    data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(df)+1, max_col=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.scaling.min = 200
    chart.y_axis.scaling.max = 400
    
    ws.add_chart(chart, "F2")
    wb.save(EXCEL_FILE)

def update_today():
    """更新今天的数据"""
    today = datetime.now().strftime('%Y%m%d')
    formatted_today = f"{today[:4]}-{today[4:6]}-{today[6:8]}"
    
    print(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 更新 {formatted_today}...")
    
    # 获取股息率
    div_rate = download_dividend_rate(today)
    if div_rate is None:
        print("  股息率: 获取失败")
        return
    print(f"  股息率: {div_rate}%")
    
    # 获取国债收益率
    bond_yield = get_bond_yield(formatted_today)
    if bond_yield is None:
        print("  国债收益率: 获取失败")
        return
    print(f"  国债收益率: {bond_yield}%")
    
    # 计算溢价
    premium = div_rate * 100 - bond_yield
    print(f"  股息率溢价: {premium:.2f}%")
    
    # 更新数据
    data = load_existing_data()
    
    # 检查是否已存在
    updated = False
    for row in data:
        if row['日期'] == formatted_today:
            row['股息率1'] = div_rate
            row['10年国债收益率'] = bond_yield
            row['股息率溢价'] = premium
            updated = True
            break
    
    if not updated:
        data.append({
            '日期': formatted_today,
            '股息率1': div_rate,
            '10年国债收益率': bond_yield,
            '股息率溢价': premium
        })
    
    # 按日期排序
    data.sort(key=lambda x: x['日期'])
    
    save_data(data)
    generate_excel()
    print(f"  ✅ 已保存!")

def backfill_history(start_date, end_date):
    """补录历史数据"""
    start = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    
    current = start
    while current <= end:
        date_str = current.strftime('%Y%m%d')
        formatted = current.strftime('%Y-%m-%d')
        
        div_rate = download_dividend_rate(date_str)
        bond_yield = get_bond_yield(formatted)
        
        if div_rate and bond_yield:
            premium = div_rate * 100 - bond_yield
            print(f"{formatted}: 股息率={div_rate}%, 国债={bond_yield}%, 溢价={premium}%")
        
        current += timedelta(days=1)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        if sys.argv[1] == "--update":
            update_today()
        elif sys.argv[1] == "--backfill":
            # python update.py --backfill 2026-01-01 2026-01-31
            backfill_history(sys.argv[2], sys.argv[3])
    else:
        update_today()
