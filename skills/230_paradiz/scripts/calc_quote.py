#!/usr/bin/env python3
import argparse
import csv
import json
from datetime import date, datetime


def d(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def normalize(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def pick(header_map, *aliases):
    for a in aliases:
        if a in header_map:
            return header_map[a]
    return None


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True)
    p.add_argument("--checkin", required=True)
    p.add_argument("--checkout", required=True)
    p.add_argument("--guests", required=True, type=int)
    p.add_argument("--room")
    args = p.parse_args()

    checkin = d(args.checkin)
    checkout = d(args.checkout)
    nights = (checkout - checkin).days
    if nights <= 0:
        print(json.dumps({"ok": False, "error": "checkout должен быть позже checkin"}, ensure_ascii=False))
        return

    path_in = args.excel
    rows = []

    if path_in.lower().endswith('.csv'):
        with open(path_in, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)
    else:
        try:
            from openpyxl import load_workbook
        except Exception:
            print(json.dumps({"ok": False, "error": "Для .xlsx нужен openpyxl (или используй .csv)"}, ensure_ascii=False))
            return
        wb = load_workbook(path_in, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print(json.dumps({"ok": False, "error": "Пустой файл с ценами"}, ensure_ascii=False))
        return

    header = [normalize(str(x) if x is not None else "") for x in rows[0]]
    hm = {h: i for i, h in enumerate(header) if h}

    c_from = pick(hm, "date_from", "from", "checkin_from", "заезд_с")
    c_to = pick(hm, "date_to", "to", "checkout_to", "выезд_по")
    c_gmin = pick(hm, "guests_min", "min_guests", "гостей_от")
    c_gmax = pick(hm, "guests_max", "max_guests", "гостей_до")
    c_ppn = pick(hm, "price_per_night", "night_price", "цена_за_ночь")
    c_total = pick(hm, "total_price", "цена_итого")
    c_curr = pick(hm, "currency", "валюта")
    c_room = pick(hm, "room", "room_type", "номер")
    c_meal = pick(hm, "meal", "питание")

    required = [c_from, c_to]
    if any(x is None for x in required) or (c_ppn is None and c_total is None):
        print(json.dumps({
            "ok": False,
            "error": "Неверная структура Excel. Нужны date_from/date_to и price_per_night или total_price"
        }, ensure_ascii=False))
        return

    # Ищем наценку за доп. человека (можно менять в прайсе вручную строкой "Доп. человек")
    extra_per_guest_per_night = 800.0
    for r in rows[1:]:
        if not r:
            continue
        room_name = str(r[c_room]).strip() if c_room is not None and c_room < len(r) and r[c_room] is not None else ""
        if room_name.lower() != "доп. человек":
            continue
        try:
            rf = r[c_from]
            rt = r[c_to]
            if isinstance(rf, datetime):
                rf = rf.date()
            elif isinstance(rf, str):
                rf = d(rf)
            if isinstance(rt, datetime):
                rt = rt.date()
            elif isinstance(rt, str):
                rt = d(rt)
            if rf <= checkin and rt >= checkout:
                extra_per_guest_per_night = float(r[c_ppn])
                break
        except Exception:
            continue

    matches = []
    for r in rows[1:]:
        if r is None:
            continue
        try:
            rf = r[c_from]
            rt = r[c_to]
            if isinstance(rf, datetime):
                rf = rf.date()
            elif isinstance(rf, str):
                rf = d(rf)
            if isinstance(rt, datetime):
                rt = rt.date()
            elif isinstance(rt, str):
                rt = d(rt)
        except Exception:
            continue

        if not (rf <= checkin and rt >= checkout):
            continue

        gmin = int(r[c_gmin]) if c_gmin is not None and r[c_gmin] is not None else 1
        gmax = int(r[c_gmax]) if c_gmax is not None and r[c_gmax] is not None else 99
        if not (gmin <= args.guests <= gmax):
            continue

        room = str(r[c_room]).strip() if c_room is not None and r[c_room] is not None else "Стандарт"
        if room.lower() == "доп. человек":
            continue
        if args.room and room.lower() != args.room.lower():
            continue

        if c_total is not None and r[c_total] is not None:
            base_total = float(r[c_total])
        else:
            base_total = float(r[c_ppn]) * nights

        # Базовая вместимость зависит от типа номера:
        # - обычные номера: база за 2 гостей (+800 ₽/сутки за 3-го и 4-го)
        # - двухкомнатный и большой с кухней: база за 4 гостей (+800 ₽/сутки за 5-го и 6-го)
        room_key = room.strip().lower()
        base_included_guests = 4 if room_key in {"двухкомнатный номер", "номер большой с кухней"} else 2
        extra_guests = max(0, args.guests - base_included_guests)
        extra_per_night = extra_per_guest_per_night * extra_guests
        total = base_total + (extra_per_night * nights)

        matches.append({
            "room": room,
            "meal": str(r[c_meal]).strip() if c_meal is not None and r[c_meal] is not None else "без питания",
            "currency": str(r[c_curr]).strip() if c_curr is not None and r[c_curr] is not None else "₽",
            "total": round(total, 2),
            "nights": nights,
        })

    if not matches:
        print(json.dumps({"ok": True, "found": 0, "message": "Подходящих тарифов не найдено"}, ensure_ascii=False))
        return

    matches.sort(key=lambda x: x["total"])
    print(json.dumps({"ok": True, "found": len(matches), "options": matches[:3]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
