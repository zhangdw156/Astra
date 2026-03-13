#!/usr/bin/env python3
import argparse
import csv
import json
import re
import subprocess
from pathlib import Path
from urllib.parse import quote_plus


def mc_call(tool: str, args: dict, timeout: int = 20):
    p = subprocess.run(
        ["mcporter", "call", tool, "--args", json.dumps(args), "--output", "json"],
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    if p.returncode != 0 or not p.stdout.strip():
        return {}
    try:
        return json.loads(p.stdout)
    except Exception:
        return {}


def parse_args():
    ap = argparse.ArgumentParser(description="Generate leads from google-maps MCP and export CSV/XLSX")
    ap.add_argument("--queries", required=True, help="Path to text file (one query per line)")
    ap.add_argument("--country-filter", default="", help="Substring filter in formatted address (e.g., 'CountryName')")
    ap.add_argument("--limit", type=int, default=50)
    ap.add_argument("--out", required=True, help="Output file (.csv or .xlsx)")
    return ap.parse_args()


def load_queries(path: Path):
    return [ln.strip() for ln in path.read_text(encoding="utf-8").splitlines() if ln.strip() and not ln.strip().startswith("#")]


def to_xlsx(rows, out_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    headers = ["name", "address", "phone", "website", "email", "rating", "place_id", "google_maps_url"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(name="Arial", bold=True)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Arial")
    for c, w in {1: 30, 2: 44, 3: 20, 4: 36, 5: 30, 6: 8, 7: 36, 8: 64}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(out_path)


def to_csv(rows, out_path: Path):
    headers = ["name", "address", "phone", "website", "email", "rating", "place_id", "google_maps_url"]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)


def main():
    args = parse_args()
    qpath = Path(args.queries)
    out = Path(args.out)
    queries = load_queries(qpath)

    seen = {}
    country_filter = args.country_filter.lower().strip()

    for q in queries:
        data = mc_call("google-maps.maps_search_places", {"query": q})
        for pl in data.get("places", []) if isinstance(data, dict) else []:
            pid = pl.get("place_id")
            if not pid or pid in seen:
                continue
            addr = pl.get("formatted_address", "")
            if country_filter and country_filter not in addr.lower():
                continue
            seen[pid] = pl
            if len(seen) >= args.limit:
                break
        if len(seen) >= args.limit:
            break

    rows = []
    for pid, pl in list(seen.items())[: args.limit]:
        d = mc_call("google-maps.maps_place_details", {"place_id": pid})
        d = d.get("result", d) if isinstance(d, dict) else {}
        name = pl.get("name", "")
        website = d.get("website", "")
        rows.append(
            {
                "name": name,
                "address": d.get("formatted_address") or pl.get("formatted_address", ""),
                "phone": d.get("formatted_phone_number") or d.get("international_phone_number", ""),
                "website": website,
                "email": "",  # website crawl optional step; keep empty by default
                "rating": d.get("rating") if d.get("rating") is not None else pl.get("rating", ""),
                "place_id": pid,
                "google_maps_url": f"https://www.google.com/maps/search/?api=1&query={quote_plus(name)}&query_place_id={pid}",
            }
        )

    out.parent.mkdir(parents=True, exist_ok=True)
    if out.suffix.lower() == ".xlsx":
        to_xlsx(rows, out)
    else:
        to_csv(rows, out)

    print(
        json.dumps(
            {
                "count": len(rows),
                "output": str(out),
                "with_phone": sum(1 for r in rows if r.get("phone")),
                "with_website": sum(1 for r in rows if r.get("website")),
                "with_email": sum(1 for r in rows if r.get("email")),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
